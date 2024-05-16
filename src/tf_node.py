#!/usr/bin/env python

import rospy
import tf
import openpyxl
from geometry_msgs.msg import TransformStamped

class TfToXlsx:
    def __init__(self):
        rospy.init_node('tf_to_xlsx_node')

        # Create Excel workbooks and worksheets
        self.workbook_2 = openpyxl.Workbook()
        self.workbook_4 = openpyxl.Workbook()
        self.workbook_6 = openpyxl.Workbook()
        
        self.sheet_2 = self.workbook_2.active
        self.sheet_4 = self.workbook_4.active
        self.sheet_6 = self.workbook_6.active

        self.sheet_2.title = "Tag36h11:2"
        self.sheet_4.title = "Tag36h11:4"
        self.sheet_6.title = "Tag36h11:6"

        # Add headers
        self.sheet_2.append(["X", "Y", "Z"])
        self.sheet_4.append(["X", "Y", "Z"])
        self.sheet_6.append(["X", "Y", "Z"])

        # TF listener
        self.listener = tf.TransformListener()

        # Run update method in a loop
        self.rate = rospy.Rate(10.0)
        self.run()

    def run(self):
        while not rospy.is_shutdown():
            self.update_coordinates("tag36h11:2", self.sheet_2)
            self.update_coordinates("tag36h11:4", self.sheet_4)
            self.update_coordinates("tag36h11:6", self.sheet_6)
            self.rate.sleep()

    def update_coordinates(self, frame_id, sheet):
        try:
            now = rospy.Time.now()
            self.listener.waitForTransform("camera_1", frame_id, now, rospy.Duration(1.0))
            (trans, rot) = self.listener.lookupTransform("/camera_1", frame_id, now)

            # Append new row to the respective sheet
            sheet.append([trans[0], trans[1], trans[2]])

        except (tf.LookupException, tf.ConnectivityException, tf.ExtrapolationException):
            rospy.logerr("Could not get transform for %s", frame_id)

    def save_files(self):
        self.workbook_2.save("tag36h11_2.xlsx")
        self.workbook_4.save("tag36h11_4.xlsx")
        self.workbook_6.save("tag36h11_6.xlsx")

if __name__ == '__main__':
    try:
        node = TfToXlsx()
        rospy.spin()
    except rospy.ROSInterruptException:
        pass
    finally:
        node.save_files()
